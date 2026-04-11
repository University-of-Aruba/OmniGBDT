from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
import tempfile
import urllib.request
import zipfile

import numpy as np

from omnigbdt import MultiOutputGBDT, SingleOutputGBDT, Verbosity


UCI_STOCK_PORTFOLIO_MIRROR_URL = (
    "https://cdn.uci-ics-mlr-prod.aws.uci.edu/390/stock%2Bportfolio%2Bperformance.zip"
)
UCI_STOCK_PORTFOLIO_WORKBOOK_NAME = "stock portfolio performance data set.xlsx"
UCI_STOCK_PORTFOLIO_FEATURE_COLUMNS = [
    "Large B/P",
    "Large ROE",
    "Large S/P",
    "Large Return Rate in the last quarter",
    "Large Market Value",
    "Small systematic Risk",
]
UCI_STOCK_PORTFOLIO_TARGET_COLUMNS = [
    "Annual Return.1",
    "Excess Return.1",
    "Systematic Risk.1",
    "Total Risk.1",
    "Abs. Win Rate.1",
    "Rel. Win Rate.1",
]


@dataclass(frozen=True)
class DeterminismDataset:
    """Store fixed train, validation, and holdout arrays for repeatability tests."""

    x_train: np.ndarray
    y_train_single: np.ndarray
    y_train_multi: np.ndarray
    x_valid: np.ndarray
    y_valid_single: np.ndarray
    y_valid_multi: np.ndarray
    x_test: np.ndarray


def make_synthetic_determinism_dataset() -> DeterminismDataset:
    """Build the correlated synthetic regression dataset used in determinism tests.

    Returns:
        DeterminismDataset: Fixed train, validation, and holdout splits for the
        synthetic repeatability matrix.
    """
    rng = np.random.default_rng(0)
    x = rng.normal(size=(400, 6))
    shared_signal = 1.2 * x[:, 0] - 0.8 * x[:, 1] + 0.5 * x[:, 2] * x[:, 3]
    y_multi = np.column_stack(
        [
            shared_signal + 0.3 * x[:, 4],
            shared_signal - 0.2 * x[:, 5],
        ]
    ).astype(np.float64)
    y_single = np.ascontiguousarray(y_multi[:, 0], dtype=np.float64)

    return DeterminismDataset(
        x_train=np.ascontiguousarray(x[:240], dtype=np.float64),
        y_train_single=np.ascontiguousarray(y_single[:240], dtype=np.float64),
        y_train_multi=np.ascontiguousarray(y_multi[:240], dtype=np.float64),
        x_valid=np.ascontiguousarray(x[240:320], dtype=np.float64),
        y_valid_single=np.ascontiguousarray(y_single[240:320], dtype=np.float64),
        y_valid_multi=np.ascontiguousarray(y_multi[240:320], dtype=np.float64),
        x_test=np.ascontiguousarray(x[320:], dtype=np.float64),
    )


def make_realworld_determinism_dataset(frame) -> DeterminismDataset:
    """Build the UCI stock-portfolio repeatability dataset from a table object.

    Args:
        frame: Table object returned by ``fetch_ucirepo(...).data.original``.

    Returns:
        DeterminismDataset: Fixed train, validation, and holdout splits for the
        real-world repeatability matrix.
    """
    feature_columns = [
        "Large B/P",
        "Large ROE",
        "Large S/P",
        "Large Return Rate in the last quarter",
        "Large Market Value",
        "Small systematic Risk",
    ]
    target_columns = [
        "Annual Return.1",
        "Excess Return.1",
        "Systematic Risk.1",
        "Total Risk.1",
        "Abs. Win Rate.1",
        "Rel. Win Rate.1",
    ]

    x = frame.loc[:, feature_columns].to_numpy(dtype=np.float64)
    y_multi = frame.loc[:, target_columns].to_numpy(dtype=np.float64)
    y_single = np.ascontiguousarray(y_multi[:, 0], dtype=np.float64)

    rng = np.random.default_rng(0)
    indices = rng.permutation(len(x))
    train_end = int(len(x) * 0.6)
    valid_end = int(len(x) * 0.8)
    train_idx = indices[:train_end]
    valid_idx = indices[train_end:valid_end]
    test_idx = indices[valid_end:]

    return DeterminismDataset(
        x_train=np.ascontiguousarray(x[train_idx], dtype=np.float64),
        y_train_single=np.ascontiguousarray(y_single[train_idx], dtype=np.float64),
        y_train_multi=np.ascontiguousarray(y_multi[train_idx], dtype=np.float64),
        x_valid=np.ascontiguousarray(x[valid_idx], dtype=np.float64),
        y_valid_single=np.ascontiguousarray(y_single[valid_idx], dtype=np.float64),
        y_valid_multi=np.ascontiguousarray(y_multi[valid_idx], dtype=np.float64),
        x_test=np.ascontiguousarray(x[test_idx], dtype=np.float64),
    )


def _deduplicate_headers(headers: list[str]) -> list[str]:
    """Normalize workbook headers and suffix duplicate names deterministically.

    Args:
        headers: Raw header strings read from the workbook.

    Returns:
        list[str]: Normalized headers with duplicate names suffixed as ``.1``,
        ``.2``, and so on.
    """
    counts: dict[str, int] = {}
    normalized: list[str] = []
    for header in headers:
        count = counts.get(header, 0)
        if count == 0:
            normalized.append(header)
        else:
            normalized.append(f"{header}.{count}")
        counts[header] = count + 1
    return normalized


def _download_uci_stock_portfolio_zip(cache_dir: Path) -> Path:
    """Download the official UCI mirror ZIP into a local cache directory.

    Args:
        cache_dir: Cache directory used for the mirror dataset artifacts.

    Returns:
        Path: Local ZIP file path.
    """
    cache_dir.mkdir(parents=True, exist_ok=True)
    zip_path = cache_dir / "stock-portfolio-performance.zip"
    if zip_path.exists():
        return zip_path

    with urllib.request.urlopen(UCI_STOCK_PORTFOLIO_MIRROR_URL) as response:
        zip_path.write_bytes(response.read())
    return zip_path


def _extract_uci_stock_portfolio_workbook(zip_path: Path, cache_dir: Path) -> Path:
    """Extract the official UCI workbook from the cached ZIP file.

    Args:
        zip_path: Cached ZIP file path.
        cache_dir: Cache directory used for extracted artifacts.

    Returns:
        Path: Local workbook path.
    """
    workbook_path = cache_dir / UCI_STOCK_PORTFOLIO_WORKBOOK_NAME
    if workbook_path.exists():
        return workbook_path

    with zipfile.ZipFile(zip_path) as archive:
        archive.extract(UCI_STOCK_PORTFOLIO_WORKBOOK_NAME, path=cache_dir)
    return workbook_path


@lru_cache(maxsize=1)
def make_realworld_determinism_dataset_from_mirror() -> DeterminismDataset:
    """Build the real-world determinism dataset from the official UCI mirror.

    Returns:
        DeterminismDataset: Fixed train, validation, and holdout splits loaded
        from the mirrored UCI workbook.
    """
    from openpyxl import load_workbook

    cache_dir = Path(tempfile.gettempdir()) / "omnigbdt-uci-stock-portfolio"
    zip_path = _download_uci_stock_portfolio_zip(cache_dir)
    workbook_path = _extract_uci_stock_portfolio_workbook(zip_path, cache_dir)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    raw_headers = [str(value).strip() for value in rows[1]]
    headers = _deduplicate_headers(raw_headers)
    row_start = 2
    data_rows = [row for row in rows[row_start:] if row and row[0] is not None]
    column_positions = {name: index for index, name in enumerate(headers)}

    x = np.ascontiguousarray(
        [
            [float(row[column_positions[name]]) for name in UCI_STOCK_PORTFOLIO_FEATURE_COLUMNS]
            for row in data_rows
        ],
        dtype=np.float64,
    )
    y_multi = np.ascontiguousarray(
        [
            [float(row[column_positions[name]]) for name in UCI_STOCK_PORTFOLIO_TARGET_COLUMNS]
            for row in data_rows
        ],
        dtype=np.float64,
    )
    y_single = np.ascontiguousarray(y_multi[:, 0], dtype=np.float64)

    rng = np.random.default_rng(0)
    indices = rng.permutation(len(x))
    train_end = int(len(x) * 0.6)
    valid_end = int(len(x) * 0.8)
    train_idx = indices[:train_end]
    valid_idx = indices[train_end:valid_end]
    test_idx = indices[valid_end:]

    return DeterminismDataset(
        x_train=np.ascontiguousarray(x[train_idx], dtype=np.float64),
        y_train_single=np.ascontiguousarray(y_single[train_idx], dtype=np.float64),
        y_train_multi=np.ascontiguousarray(y_multi[train_idx], dtype=np.float64),
        x_valid=np.ascontiguousarray(x[valid_idx], dtype=np.float64),
        y_valid_single=np.ascontiguousarray(y_single[valid_idx], dtype=np.float64),
        y_valid_multi=np.ascontiguousarray(y_multi[valid_idx], dtype=np.float64),
        x_test=np.ascontiguousarray(x[test_idx], dtype=np.float64),
    )


def make_determinism_params(num_threads: int, early_stop: int) -> dict[str, object]:
    """Build a shared parameter dictionary for deterministic repeatability tests.

    Args:
        num_threads: Requested native thread count.
        early_stop: Early-stopping patience.

    Returns:
        dict[str, object]: Training parameter dictionary for deterministic tests.
    """
    return {
        "loss": b"mse",
        "max_depth": 4,
        "max_bins": 128,
        "lr": 0.05,
        "deterministic": True,
        "seed": 23,
        "num_threads": num_threads,
        "early_stop": early_stop,
        "verbosity": Verbosity.SILENT,
    }


def train_single_determinism_run(
    dataset: DeterminismDataset,
    num_threads: int,
    early_stop: int,
    num_rounds: int,
    model_path: Path,
) -> tuple[bytes, np.ndarray]:
    """Train one single-output repeatability run and collect its artifacts.

    Args:
        dataset: Fixed dataset splits for the run.
        num_threads: Requested native thread count.
        early_stop: Early-stopping patience.
        num_rounds: Maximum number of boosting rounds.
        model_path: Output path for the dumped model.

    Returns:
        tuple[bytes, numpy.ndarray]: Dumped model bytes and holdout predictions.
    """
    booster = SingleOutputGBDT(params=make_determinism_params(num_threads, early_stop))
    try:
        booster.set_data(
            (dataset.x_train, dataset.y_train_single),
            (dataset.x_valid, dataset.y_valid_single),
        )
        booster.train(num_rounds)
        predictions = np.ascontiguousarray(booster.predict(dataset.x_test))
        booster.dump(model_path)
        return model_path.read_bytes(), predictions
    finally:
        booster.close()


def train_multi_determinism_run(
    dataset: DeterminismDataset,
    num_threads: int,
    early_stop: int,
    num_rounds: int,
    model_path: Path,
) -> tuple[bytes, np.ndarray]:
    """Train one multi-output repeatability run and collect its artifacts.

    Args:
        dataset: Fixed dataset splits for the run.
        num_threads: Requested native thread count.
        early_stop: Early-stopping patience.
        num_rounds: Maximum number of boosting rounds.
        model_path: Output path for the dumped model.

    Returns:
        tuple[bytes, numpy.ndarray]: Dumped model bytes and holdout predictions.
    """
    booster = MultiOutputGBDT(
        out_dim=dataset.y_train_multi.shape[1],
        params=make_determinism_params(num_threads, early_stop),
    )
    try:
        booster.set_data(
            (dataset.x_train, dataset.y_train_multi),
            (dataset.x_valid, dataset.y_valid_multi),
        )
        booster.train(num_rounds)
        predictions = np.ascontiguousarray(booster.predict(dataset.x_test))
        booster.dump(model_path)
        return model_path.read_bytes(), predictions
    finally:
        booster.close()


def assert_repeated_runs_are_deterministic(
    train_once,
    tmp_path: Path,
    repeats: int = 3,
) -> None:
    """Assert byte-for-byte and prediction determinism across repeated runs.

    Args:
        train_once: Callable that accepts a model path and returns dumped model
            bytes together with holdout predictions.
        tmp_path: Temporary directory for model dumps.
        repeats: Number of repeated training runs.
    """
    reference_dump = None
    reference_predictions = None

    for repeat in range(repeats):
        dump_bytes, predictions = train_once(tmp_path / f"repeat_{repeat}.txt")
        if reference_dump is None:
            reference_dump = dump_bytes
            reference_predictions = predictions
            continue

        assert dump_bytes == reference_dump
        assert np.array_equal(predictions, reference_predictions)
